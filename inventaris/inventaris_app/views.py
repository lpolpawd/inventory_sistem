from django.shortcuts import render, redirect
from .models import Item,  Report, Transaction 
from .forms import ItemForm, TransactionForm
import json
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from django.core.serializers.json import DjangoJSONEncoder
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User, Group 
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required, permission_required
from .utils import get_inventory_summary
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from django.http import HttpResponse, JsonResponse
from django.utils import timezone  # Import timezone
from django.db.models import Q, Sum, Count
from django.db import models

#start of inventaris logic
@permission_required('inventaris_app.add_item', raise_exception=True)
def add_item(request):
    if request.method == 'POST':
        form = ItemForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('view_inventory')
    else:
        form = ItemForm()
    return render(request, 'inventaris_app/add_item.html', {'form': form})

def update_item(request, item_id):
    item = Item.objects.get(id=item_id)
    if request.method == 'POST':
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('view_inventory')
    else:
        form = ItemForm(instance=item)
    return render(request, 'inventaris_app/update_item.html', {'form': form})

@permission_required('inventaris_app.delete_item', raise_exception=True)
def delete_item(request, item_id):
    item = Item.objects.get(id=item_id)
    if request.method == 'POST':
        item.delete()
        return redirect('view_inventory')
    return render(request, 'inventaris_app/delete_item.html', {'item': item})

def add_transaction(request):
    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save()
            item = transaction.item
            if transaction.transaction_type == 'ADD':
                item.stock += transaction.quantity
            else:
                item.stock -= transaction.quantity
            item.save()
            return redirect('view_inventory')
    else:
        form = TransactionForm()
    return render(request, 'inventaris_app/add_transaction.html', {'form': form})

@login_required
def view_inventory(request):
    query = request.GET.get('q')
    filter_by = request.GET.get('filter_by')

    items = Item.objects.all()

    if query:
        items = items.filter(
            Q(name__icontains=query) | Q(description__icontains=query)
        )

    if filter_by == 'low_stock':
        items = items.filter(stock__lt=10)
    elif filter_by == 'high_stock':
        items = items.filter(stock__gte=10)
    return render(request, 'inventaris_app/view_inventory.html', {'items': items})

class DateTimeEncoder(DjangoJSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        return super().default(obj)

def generate_report(request):
    # Get the inventory summary using the helper function
    report_content = get_inventory_summary()

    # Store the report (if needed)
    report = Report.objects.create(content=json.dumps(report_content, cls=DjangoJSONEncoder))

    return render(request, 'inventaris_app/report.html', {'report': report_content})
#End of inventaris logic

#start of user autentikasi
def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')  # Redirect to login page after successful registration
    else:
        form = UserCreationForm()
    return render(request, 'inventaris_app/register.html', {'form': form})

def user_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('view_inventory')  # Redirect to a protected view after login
    else:
        form = AuthenticationForm()
    return render(request, 'inventaris_app/login.html', {'form': form})

def user_logout(request):
    logout(request)
    return redirect('login')

def assign_user_to_group(user_id, group_name):
    user = User.objects.get(id=user_id)
    group = Group.objects.get(name=group_name)
    user.groups.add(group)

def base(request):
    return render(request, 'inventaris_app/base.html')

def inventory_dashboard(request):
    #  Fetch data
    items = Item.objects.all()
    low_stock_items = Item.objects.filter(stock__lt=10)
    today = datetime.now().date()
    past_30_days = today - timedelta(days=30)
    
    # Convert QuerySet to list of dictionaries
    stock_trends = (
        Transaction.objects.filter(date__gte=past_30_days)
        .values('item__name')
        .annotate(
            total_added=Sum('quantity', filter=Q(transaction_type='ADD')),
            total_removed=Sum('quantity', filter=Q(transaction_type='REMOVE')),
            transaction_count=Count('id')
        )
    )
    
    # Convert QuerySet to list and replace None with 0
    stock_trends_list = list(stock_trends)
    for trend in stock_trends_list:
        trend['total_added'] = trend.get('total_added', 0) or 0
        trend['total_removed'] = trend.get('total_removed', 0) or 0

    # Serialize to JSON
    stock_trends_json = json.dumps(stock_trends_list)

    context = {
        'items': items,
        'low_stock_items': low_stock_items,
        'stock_trends_json': stock_trends_json,  # Pass JSON data
    }
    return render(request, 'inventaris_app/dashboard.html', context)

def export_report_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inventory_report.pdf"'

    # Create a SimpleDocTemplate object with the response object and set page size and margins
    doc = SimpleDocTemplate(response, pagesize=letter,
                            rightMargin=inch, leftMargin=inch,
                            topMargin=inch, bottomMargin=inch)

    elements = []

    # Title and date
    styles = getSampleStyleSheet()
    title = Paragraph("Inventory Report", styles['Title'])
    date = Paragraph(f"Date: {timezone.now().strftime('%Y-%m-%d')}", styles['Normal'])
    elements.append(title)
    elements.append(date)

    elements.append(Paragraph("<br/><br/>", styles['Normal']))

    # Create table data
    items = Item.objects.all()
    data = [['Item Name', 'Stock', 'Description']]  # Header row

    for item in items:
        data.append([item.name, item.stock, item.description])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)

    doc.build(elements)

    return response

def export_report_excel(request):
    # Create an in-memory workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Inventory Report"

    # Set up the headers
    headers = ['Item Name', 'Stock', 'Description']
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        sheet[f'{column_letter}1'] = header

    # Fetch data to include in the Excel file
    items = Item.objects.all()
    for row_num, item in enumerate(items, 2):
        sheet[f'A{row_num}'] = item.name
        sheet[f'B{row_num}'] = item.stock
        sheet[f'C{row_num}'] = item.description

    # Create an HttpResponse object with Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=inventory_report.xlsx'
    workbook.save(response)

    return response